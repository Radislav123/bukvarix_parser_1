import glob
import os
import random
import shutil
import time

import openpyxl
import pandas
from selenium.webdriver import Chrome, ChromeOptions
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from pages import LoginPage, SearchPage
from . import models, settings
from .secret_keeper import SecretKeeper


class BukvarixParser:
    """Отвечает за весь процесс парсинга."""

    progress: models.Progress
    driver: Chrome
    secrets: SecretKeeper
    _proxies: dict = None

    def setup_method(self) -> None:
        if os.path.exists(settings.DOWNLOADS):
            shutil.rmtree(settings.DOWNLOADS)
        self.secrets = SecretKeeper()

        options = ChromeOptions()
        # этот параметр тоже нужен, так как в режиме headless с некоторыми элементами нельзя взаимодействовать
        options.add_argument("--disable-blink-features=AutomationControlled")
        # todo: return headless mode
        # options.add_argument("--headless")
        # options.add_argument("--window-size=1920,1080")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])

        prefs = {"download.default_directory": settings.DOWNLOADS}
        options.add_experimental_option("prefs", prefs)

        driver_manager = ChromeDriverManager(path = "").install()
        service = Service(executable_path = driver_manager)

        parameters = {
            "options": options,
            "service": service
        }
        self.driver = Chrome(**parameters)
        self.driver.maximize_window()

    def teardown_method(self) -> None:
        self.driver.quit()

    def update_progress(self, addition: int) -> None:
        self.progress.current += addition
        self.progress.save()

    @staticmethod
    def get_domains() -> list[str]:
        # todo: заменить TEST_DOMAINS на DOMAINS
        book = openpyxl.load_workbook(settings.TEST_DOMAINS)
        sheet = book.active
        domains = []
        row = 1
        while sheet.cell(row, 1).value:
            domains.append(sheet.cell(row, 1).value)
            row += 1
        return domains

    @staticmethod
    def get_filtered_domains(dataframe: pandas.DataFrame) -> dict[str, tuple[str, pandas.DataFrame]]:
        """Возвращает список доменов, количество ключевых слов которого превышает 50 (settings.DOMAIN_WORDS_AMOUNT)."""

        domains = {}
        for header in [x for x in list(dataframe) if "Позиция в Google" in x]:
            header: str
            domain = header.split()[-1]
            filtered = dataframe[dataframe[header] > 0]
            if len(filtered) > settings.DOMAIN_WORDS_AMOUNT:
                domains[header] = (domain, filtered)

        return domains

    def process_downloaded_data(self) -> None:
        files = glob.glob(f"{settings.DOWNLOADS}/*")
        for filename in files:
            dataframe = pandas.read_csv(filename, delimiter = ';')
            filtered_domains = self.get_filtered_domains(dataframe)
            for header in filtered_domains:
                domain = models.Domain(name = filtered_domains[header][0])
                filtered = filtered_domains[header][1]
                domain.average_position = int(filtered[header].mean())
                top_10 = filtered[filtered[header] <= 10]
                top_3 = filtered[filtered[header] <= 3]
                domain.requests_top_10 = len(top_10)
                domain.requests_top_3 = len(top_3)
                domain.frequency_sum_top_10 = top_10['"!Частотность !Весь !мир"'].sum()
                domain.frequency_sum_top_3 = top_3['"!Частотность !Весь !мир"'].sum()
                domain.save()
            self.update_progress(settings.REQUEST_DOMAINS_AMOUNT)
        self.update_progress(self.progress.capacity - self.progress.current)

    def run(self) -> None:
        login_page = LoginPage(self.driver)
        login_page.open()
        login_page.login(self.secrets.bukvarix.login, self.secrets.bukvarix.password)

        domains = self.get_domains()
        progress_capacity = len(domains) * 2
        self.progress = models.Progress(capacity = progress_capacity)
        self.progress.save()

        for start in range(0, len(domains), settings.REQUEST_DOMAINS_AMOUNT):
            search_page = SearchPage(self.driver)
            search_page.open()
            search_page.search(domains[start:start + settings.REQUEST_DOMAINS_AMOUNT])
            search_page.download_button.click()
            addition = settings.REQUEST_DOMAINS_AMOUNT
            if start + addition > len(domains):
                addition = len(domains) - start
            self.update_progress(addition)
            time.sleep(random.randint(15, 45))

        # время, чтобы скачанные файлы сохранились правильно,
        # иначе последний файл сохраняется как временный и может вообще не успеть скачаться
        time.sleep(3)

        self.process_downloaded_data()
